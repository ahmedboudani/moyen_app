from models import db, Class, Student, Grade, GeneralInfo, User
import os
from export_pdf import export_to_pdf
import pandas as pd
from werkzeug.utils import secure_filename
from flask import Flask, render_template, request, send_file, jsonify,flash , redirect,url_for
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import check_password_hash
from openpyxl import load_workbook # ✨ المكتبة الجديدة         # ✨ للتعامل مع الملف في الذا
import io
from openpyxl.utils import coordinate_to_tuple
import os
from werkzeug.utils import secure_filename
import openpyxl
# Ensure the instance folder exists
instance_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance')
os.makedirs(instance_path, exist_ok=True)

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///instance/test.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db.init_app(app)


# تعديل رابط قاعدة البيانات إذا كان من Render
if app.config['SQLALCHEMY_DATABASE_URI'] and app.config['SQLALCHEMY_DATABASE_URI'].startswith('postgres://'):
    app.config['SQLALCHEMY_DATABASE_URI'] = app.config['SQLALCHEMY_DATABASE_URI'].replace('postgres://', 'postgresql://', 1)



login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@app.route('/db-test')
def db_test():
    try:
        # محاولة إجراء استعلام بسيط
        users_count = User.query.count()
        return f"تم الاتصال بقاعدة البيانات بنجاح! عدد المستخدمين: {users_count}"
    except Exception as e:
        return f"خطأ في الاتصال بقاعدة البيانات: {str(e)}"

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# ✅ توحيد أسماء الفصول
ALL_SEMESTERS = ["الفصل الأول", "الفصل الثاني", "الفصل الثالث"]
ALL_SUBJECTS = [
    "الرياضيات", "اللغة العربية", "اللغة الفرنسية", "اللغة الإنجليزية",
    "العلوم الطبيعية", "العلوم الفيزيائية", "التاريخ", "الجغرافيا",
    "التربية الإسلامية", "التربية المدنية", "التربية البدنية",
    "المعلوماتية", "الفنون التشكيلية", "الموسيقى"
]

@app.context_processor
def inject_general_info():
    if current_user.is_authenticated:
        info = GeneralInfo.query.filter_by(user_id=current_user.id).first()
        return dict(general_info=info)
    return dict(general_info=None)

# إنشاء جداول قاعدة البيانات
with app.app_context():
    db.create_all()

# ----------------- Authentication -----------------
@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('classes'))
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = User.query.filter_by(username=username).first()
        if user:
            flash('Username already exists.')
            return redirect(url_for('register'))
        try:
            new_user = User(username=username)
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.commit()
            flash('Congratulations, you are now a registered user!')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            print(f"Error during registration: {e}")
            flash(f"An error occurred during registration: {e}", "error")
            return redirect(url_for('register'))
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    # 1. إذا كان المستخدم مسجلاً بالفعل، أعد توجيهه إلى صفحة الأقسام
    if current_user.is_authenticated:
        return redirect(url_for('classes'))

    # 2. معالجة طلب POST (محاولة تسجيل الدخول)
    if request.method == 'POST':
        try:
            username = request.form.get('username')
            password = request.form.get('password')
            
            # البحث عن المستخدم
            user = User.query.filter_by(username=username).first()

            # التحقق من وجود المستخدم وصحة كلمة المرور
            if user and user.check_password(password):
                login_user(user)
                
                # التوجيه إلى الصفحة المطلوبة أو إلى صفحة الأقسام الافتراضية
                next_page = request.args.get('next')
                return redirect(next_page or url_for('classes'))
            else:
                # رسالة خطأ عند فشل تسجيل الدخول
                flash('اسم المستخدم أو كلمة المرور غير صحيحة.', 'error')
                # ابقَ على صفحة تسجيل الدخول لتمكين المستخدم من المحاولة مرة أخرى
                return render_template('login.html')

        except Exception as e:
            # التعامل مع الأخطاء غير المتوقعة أثناء عملية تسجيل الدخول
            print(f"Error during login POST: {e}")
            flash("حدث خطأ داخلي أثناء محاولة تسجيل الدخول. يرجى المحاولة مرة أخرى.", "error")
            return redirect(url_for('login'))
            
    # 3. معالجة طلب GET (عرض نموذج تسجيل الدخول)
    return render_template('login.html')
@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))

# ---------------- الصفحة الرئيسية ----------------
@app.route("/", methods=["GET", "POST"])
@login_required
def index():
    info = GeneralInfo.query.filter_by(user_id=current_user.id).first()

    if request.method == "POST":
        institution = request.form.get("institution")
        directorate = request.form.get("directorate")
        teacher = request.form.get("teacher")
        subject = request.form.get("subject")

        if info:
            info.institution = institution
            info.directorate = directorate
            info.teacher = teacher
            info.subject = subject
        else:
            info = GeneralInfo(
                institution=institution,
                directorate=directorate,
                teacher=teacher,
                subject=subject,
                user_id=current_user.id
            )
            db.session.add(info)
        db.session.commit()

    return render_template("index.html", info=info, subjects=ALL_SUBJECTS)


@app.route("/delete_info", methods=["POST"])
@login_required
def delete_info():
    GeneralInfo.query.filter_by(user_id=current_user.id).delete()
    db.session.commit()
    return redirect(url_for("index"))

# ---------------- إدارة الأقسام ----------------
@app.route("/classes", methods=["GET", "POST"])
@login_required
def classes():
    if request.method == "POST":
        name = request.form["name"]
        if name:
            db.session.add(Class(name=name, user_id=current_user.id))
            db.session.commit()
        return redirect(url_for("classes"))
    user_classes = Class.query.filter_by(user_id=current_user.id).all()
    return render_template("classes.html", classes=user_classes)

@app.route("/class/<int:id>/delete")
@login_required
def delete_class(id):
    c = Class.query.get_or_404(id)
    if c.user_id != current_user.id:
        return redirect(url_for('classes'))
    db.session.delete(c)
    db.session.commit()
    return redirect(url_for("classes"))

@app.route("/class/<int:id>/edit", methods=["POST"])
@login_required
def edit_class(id):
    c = Class.query.get_or_404(id)
    if c.user_id != current_user.id:
        return redirect(url_for('classes'))
    c.name = request.form["name"]
    db.session.commit()
    return redirect(url_for("classes"))

# ---------------- إدارة التلاميذ ----------------
@app.route("/students", methods=["GET", "POST"])
@login_required
def students():
    if request.method == "POST":
        fullname = request.form["fullname"]
        class_id = request.form["class_id"]

        if fullname and class_id:
            try:
                for sem in ALL_SEMESTERS:
                    s = Student(fullname=fullname, class_id=class_id, semester=sem, user_id=current_user.id)
                    db.session.add(s)
                db.session.commit()
            except Exception:
                db.session.rollback()
        return redirect(url_for("students", semester=request.args.get('semester', ALL_SEMESTERS[0])))

    current_semester = request.args.get('semester', ALL_SEMESTERS[0])
    user_students = Student.query.filter_by(user_id=current_user.id, semester=current_semester).order_by(Student.fullname).all()
    user_classes = Class.query.filter_by(user_id=current_user.id).all()

    return render_template(
        "students.html",
        students=user_students,
        classes=user_classes,
        semesters=ALL_SEMESTERS,
        current_semester=current_semester
    )

@app.route("/student/<int:id>/delete")
@login_required
def delete_student(id):
    student = Student.query.get_or_404(id)
    if student.user_id != current_user.id:
        return redirect(url_for('students'))
    semester = student.semester
    db.session.delete(student)
    db.session.commit()
    return redirect(url_for("students", semester=semester))

@app.route("/student/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
def edit_student(student_id):
    student = Student.query.get_or_404(student_id)
    if student.user_id != current_user.id:
        return redirect(url_for('students'))
    if request.method == "POST":
        student.fullname = request.form["fullname"]
        student.class_id = request.form["class_id"]
        student.semester = request.form["semester"]
        db.session.commit()
        return redirect(url_for("students", semester=student.semester))
    
    user_classes = Class.query.filter_by(user_id=current_user.id).all()
    return render_template("edit_student.html", student=student, classes=user_classes, semesters=ALL_SEMESTERS)

# ---------------- إدارة العلامات ----------------
@app.route("/grades/<int:student_id>", methods=["GET", "POST"])
@login_required
def grades(student_id):
    student = Student.query.get_or_404(student_id)
    if student.user_id != current_user.id:
        return redirect(url_for('students'))
    if request.method == "POST":
        g = Grade(
            assessment=float(request.form["assessment"]),
            test=float(request.form["test"]),
            exam=float(request.form["exam"]),
            student_id=student.id,
            semester=request.form["semester"],
            user_id=current_user.id
        )
        db.session.add(g)
        db.session.commit()
        return redirect(url_for("grades", student_id=student.id))

    return render_template("grades.html", student=student, grades=student.grades, semesters=ALL_SEMESTERS)

@app.route("/grade/<int:grade_id>/edit", methods=["GET", "POST"])
@login_required
def edit_grade(grade_id):
    grade = Grade.query.get_or_404(grade_id)
    student = grade.student
    if student.user_id != current_user.id:
        return redirect(url_for('students'))
    if request.method == "POST":
        grade.assessment = float(request.form["assessment"])
        grade.test = float(request.form["test"])
        grade.exam = float(request.form["exam"])
        db.session.commit()
        return redirect(url_for("grades", student_id=student.id))
    return render_template("edit_grade.html", grade=grade, student=student)

# ---------------- عرض النتائج ----------------
@app.route("/results")
@login_required
def results():
    class_id = request.args.get("class_id")
    semester = request.args.get("semester")
    query = Student.query.filter_by(user_id=current_user.id)

    if class_id and class_id.isdigit():
        query = query.filter(Student.class_id == int(class_id))
    if semester:
        query = query.filter(Student.semester == semester)

    students_with_grades = []
    all_students = query.order_by(Student.fullname).all()
    for student in all_students:
        grades_for_semester = [g for g in student.grades if g.semester == semester]
        student.grades_for_semester = grades_for_semester
        students_with_grades.append(student)

    stats = calculate_statistics(students_with_grades, semester)
    user_classes = Class.query.filter_by(user_id=current_user.id).all()

    return render_template(
        "results.html",
        students=students_with_grades,
        classes=user_classes,
        selected_class=int(class_id) if class_id and class_id.isdigit() else None,
        selected_semester=semester,
        semesters=ALL_SEMESTERS,
        num_students=stats['num_students'],
        class_average=stats['class_average'],
        success_percentage=stats['success_percentage']
    )
def calculate_statistics(students, semester):
    num_students = len(students)
    total_student_averages = 0
    successful_students = 0
    counted_students = 0

    for student in students:
        student_grades_for_semester = student.grades_for_semester
        if not student_grades_for_semester:
            continue

        student_total = sum(g.average for g in student_grades_for_semester)
        student_average = student_total / len(student_grades_for_semester)

        total_student_averages += student_average
        counted_students += 1

        if student_average >= 10:
            successful_students += 1

    class_average = total_student_averages / counted_students if counted_students else 0
    success_percentage = (successful_students / counted_students) * 100 if counted_students else 0

    return {
        'num_students': num_students,
        'counted_students': counted_students,
        'class_average': class_average,
        'success_percentage': success_percentage
    }


@app.route("/export/pdf")
@login_required
def export_pdf_route():
    class_id = request.args.get("class_id")
    semester = request.args.get("semester")

    query = Student.query.filter_by(user_id=current_user.id)
    class_name = None
    if class_id and class_id.isdigit():
        query = query.filter(Student.class_id == int(class_id))
        class_obj = Class.query.get(int(class_id))
        if class_obj and class_obj.user_id == current_user.id:
            class_name = class_obj.name

    if semester:
        query = query.filter(Student.semester == semester)

    students_with_grades = []
    all_students = query.all()
    for student in all_students:
        grades_for_semester = [g for g in student.grades if g.semester == semester]
        student.grades_for_semester = grades_for_semester
        students_with_grades.append(student)

    stats = calculate_statistics(students_with_grades, semester)
    general_info = GeneralInfo.query.filter_by(user_id=current_user.id).first()

    pdf_filename = "report.pdf"
    export_to_pdf(
        students_with_grades,
        class_name,
        semester,
        stats['num_students'],
        stats['class_average'],
        stats['success_percentage'],
        info=general_info,
        filename=pdf_filename
    )

    return send_file(pdf_filename, as_attachment=True)

import tempfile
@app.route("/import_excel", methods=["POST"])
@login_required
def import_excel():
    file = request.files.get("file")
    if not file:
        flash("لم يتم اختيار ملف", "error")
        return redirect(url_for("students"))

    sheet_name = request.form.get("sheet_name")
    start_row = request.form.get("start_row", "3")
    try:
        start_row = int(start_row) - 1
        if start_row < 0:
            start_row = 0
    except ValueError:
        start_row = 2
    
    class_id = request.form.get("class_id")
    
    try:
        selected_class = None
        if class_id and class_id.isdigit():
            selected_class = Class.query.get(int(class_id))
            if selected_class.user_id != current_user.id:
                flash("Invalid class selected.")
                return redirect(url_for('students'))

        if not selected_class:
            user_classes = Class.query.filter_by(user_id=current_user.id).all()
            if not user_classes:
                selected_class = Class(name="القسم الافتراضي", user_id=current_user.id)
                db.session.add(selected_class)
                db.session.commit()
            else:
                selected_class = user_classes[0]

        sheet_index = 0
        if sheet_name == "Sheet1" or sheet_name == "ورقة1":
            sheet_index = 0
        elif sheet_name == "Sheet2" or sheet_name == "ورقة2":
            sheet_index = 1
        elif sheet_name == "Sheet3" or sheet_name == "ورقة3":
            sheet_index = 2

        df = pd.read_excel(file, header=None, sheet_name=sheet_index)

        if df.shape[0] < start_row + 1 or df.shape[1] < 3:
            flash(f"بنية الملف غير صحيحة. يجب أن يحتوي الملف على بيانات بدءًا من السطر {start_row + 1} وعلى الأقل 3 أعمدة.", "error")
            return redirect(url_for("students"))

        students_added = 0

        for index, row in df.iloc[start_row:].iterrows():
            try:
                lastname = str(row.get(1, "")).strip()
                firstname = str(row.get(2, "")).strip()

                if not lastname or lastname.lower() == 'nan' or not firstname or firstname.lower() == 'nan':
                    continue
                    
                fullname = f"{lastname} {firstname}"
                
                for sem in ALL_SEMESTERS:
                    exists = Student.query.filter_by(
                        fullname=fullname,
                        class_id=selected_class.id,
                        semester=sem,
                        user_id=current_user.id
                    ).first()
                    
                    if not exists:
                        student = Student(
                            fullname=fullname,
                            class_id=selected_class.id,
                            semester=sem,
                            user_id=current_user.id
                        )
                        db.session.add(student)
                        students_added += 1
            
            except Exception as row_error:
                continue
        
        db.session.commit()
        
        if students_added > 0:
            flash(f"تم استيراد {students_added} طالب بنجاح من الصفحة {sheet_name} إلى القسم {selected_class.name}", "success")
        else:
            flash("لم يتم إضافة أي طالب جديد", "info")
        
        return redirect(url_for("students"))
            
    except Exception as e:
        db.session.rollback()
        flash(f"حدث خطأ أثناء استيراد الملف: {str(e)}", "error")
        return redirect(url_for("students"))

def get_top_left_cell(sheet, cell_ref):
    for merged_range in sheet.merged_cells.ranges:
        if cell_ref in merged_range:
            return merged_range.min_row, merged_range.min_col
    return coordinate_to_tuple(cell_ref)

import sqlite3
import json
@app.route("/export_excel", methods=["POST"])
@login_required
def export_excel():
    try:
        if 'excel_file' not in request.files:
            return jsonify({"success": False, "error": "لم يتم تحديد ملف"})
        
        excel_file = request.files['excel_file']
        if excel_file.filename == '':
            return jsonify({"success": False, "error": "لم يتم اختيار ملف"})
            
        sheet_name = request.form.get('sheet_name')
        students_data_json = request.form.get('students_data')
        
        if not sheet_name or not students_data_json:
            return jsonify({"success": False, "error": "بيانات غير مكتملة"})
        
        students_data = json.loads(students_data_json)
        
        file_data = excel_file.read()
        input_buffer = io.BytesIO(file_data)
        
        try:
            workbook = load_workbook(input_buffer)
            
            sheet_index = int(sheet_name) - 1
            if sheet_index < 0 or sheet_index >= len(workbook.sheetnames):
                return jsonify({"success": False, "error": f"الصفحة رقم {sheet_name} غير موجودة"})
            
            sheet = workbook.worksheets[sheet_index]
            
            start_row = 9
            
            for i, student in enumerate(students_data):
                row = start_row + i
                
                sheet[f'E{row}'] = student.get('assessment', '')
                sheet[f'F{row}'] = student.get('test', '')
        
                sheet[f'G{row}'] = student.get('exam', '')
                sheet[f'H{row}'] = student.get('note', '')
            
            output_buffer = io.BytesIO()
            workbook.save(output_buffer)
            output_buffer.seek(0)
            
            return send_file(
                output_buffer,
                as_attachment=True,
                download_name=excel_file.filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            print(f"Error processing Excel file: {str(e)}")
            return jsonify({"success": False, "error": str(e)})
            
    except Exception as e:
        print(f"General error: {str(e)}")
        return jsonify({"success": False, "error": str(e)})


if __name__ == "__main__":
    from waitress import serve
    port = int(os.environ.get("PORT", 5000))
    # Check if running on Heroku
    if 'DYNO' in os.environ:
        # Running on Heroku, use waitress
        serve(app, host='0.0.0.0', port=port)
    else:
        # Running locally
        app.run(debug=True, host='0.0.0.0', port=port)
